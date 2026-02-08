"""Stage 1: Parse Excel/PDF files to Markdown.

Uses Docling (local, no API key needed) as the primary parser.
Falls back to pandas/openpyxl for Excel files if Docling fails.

For .xlsm files, formulas are resolved to values before parsing so that
macro-enabled workbooks display their calculated data correctly.
"""

import re
import tempfile
from pathlib import Path

from config.settings import SUPPORTED_PARSE_EXTENSIONS


# ---------------------------------------------------------------------------
# Excel formula resolver (xlsm pre-processing)
# ---------------------------------------------------------------------------

def _resolve_xlsm_formulas(file_path: Path, log_callback=None) -> Path:
    """Open an xlsm workbook, resolve formulas to values, save as xlsx.

    Returns the path to a temporary .xlsx file with all values baked in.
    """
    import openpyxl

    def log(msg):
        if log_callback:
            log_callback(msg)

    log(f"Resolving formulas in {file_path.name}...")

    # Load twice: once for formulas, once for cached values
    wb_formulas = openpyxl.load_workbook(
        str(file_path), data_only=False, keep_vba=True
    )
    wb_cached = openpyxl.load_workbook(
        str(file_path), data_only=True, keep_vba=True
    )

    # Build a lookup of all cell values across all sheets (for formula resolution)
    cell_values = {}  # {('Sheet Name', 'A1'): value}
    for sheet_name in wb_cached.sheetnames:
        ws = wb_cached[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_values[(sheet_name, cell.coordinate)] = cell.value

    # Also index by (sheet_name, row, col) for column-letter-free lookups
    for sheet_name in wb_formulas.sheetnames:
        ws = wb_formulas[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is not None and not (isinstance(val, str) and val.startswith('=')):
                    cell_values[(sheet_name, cell.coordinate)] = val

    def _resolve_cell_ref(sheet, coord):
        """Look up a cell value from the cache."""
        return cell_values.get((sheet, coord))

    def _parse_cell_ref(ref_str, current_sheet):
        """Parse 'Sheet!A1' or 'A1' into (sheet_name, coordinate)."""
        ref_str = ref_str.strip().strip("'").strip('"')
        if '!' in ref_str:
            parts = ref_str.split('!', 1)
            sheet = parts[0].strip("'").strip('"')
            coord = parts[1].strip().replace('$', '')
            return sheet, coord
        return current_sheet, ref_str.strip().replace('$', '')

    def _col_letter_to_num(col_str):
        """Convert column letter(s) to 1-based number. A=1, B=2, ..., Z=26, AA=27."""
        col_str = col_str.upper().replace('$', '')
        result = 0
        for ch in col_str:
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result

    def _num_to_col_letter(n):
        """Convert 1-based column number to letter(s)."""
        result = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _parse_coord(coord_str):
        """Parse 'A1' into (col_letter, row_number)."""
        coord_str = coord_str.replace('$', '')
        m = re.match(r'^([A-Z]+)(\d+)$', coord_str, re.IGNORECASE)
        if m:
            return m.group(1).upper(), int(m.group(2))
        return None, None

    def _get_range_value(sheet, range_str, row_idx, col_idx):
        """Get a value from a sheet range by row/col index (1-based).

        range_str like 'B2:I7' — row_idx=1 means first row of range (row 2).
        """
        range_str = range_str.replace('$', '')
        m = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str, re.IGNORECASE)
        if not m:
            return None
        start_col = _col_letter_to_num(m.group(1))
        start_row = int(m.group(2))
        target_col = start_col + col_idx - 1
        target_row = start_row + row_idx - 1
        coord = f"{_num_to_col_letter(target_col)}{target_row}"
        return _resolve_cell_ref(sheet, coord)

    def _match_in_range(lookup_val, sheet, range_str, match_type=0):
        """Simulate Excel MATCH: find lookup_val in a 1D range, return 1-based position."""
        range_str = range_str.replace('$', '')
        m = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', range_str, re.IGNORECASE)
        if not m:
            return None
        start_col = _col_letter_to_num(m.group(1))
        start_row = int(m.group(2))
        end_col = _col_letter_to_num(m.group(3))
        end_row = int(m.group(4))

        if start_col == end_col:
            # Vertical range (column)
            for i, r in enumerate(range(start_row, end_row + 1), 1):
                coord = f"{_num_to_col_letter(start_col)}{r}"
                val = _resolve_cell_ref(sheet, coord)
                if val is not None and str(val).strip() == str(lookup_val).strip():
                    return i
        elif start_row == end_row:
            # Horizontal range (row)
            for i, c in enumerate(range(start_col, end_col + 1), 1):
                coord = f"{_num_to_col_letter(c)}{start_row}"
                val = _resolve_cell_ref(sheet, coord)
                if val is not None and str(val).strip() == str(lookup_val).strip():
                    return i
        return None

    def _resolve_ref_or_value(expr, current_sheet):
        """Resolve expression: either a cell ref, VLOOKUP, or a literal value."""
        expr = expr.strip()
        if expr.startswith('"') and expr.endswith('"'):
            return expr.strip('"')
        # Handle VLOOKUP inside expressions
        m_vl = re.match(
            r"VLOOKUP\(\s*(.+?)\s*,\s*(.+?\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)\s*,\s*(\d+)\s*,\s*(?:FALSE|0)\s*\)",
            expr, re.IGNORECASE
        )
        if m_vl:
            return _resolve_vlookup(m_vl.group(1), m_vl.group(2), int(m_vl.group(3)), current_sheet)
        # Check if it's a cell reference
        clean = expr.replace('$', '').replace("'", '')
        if re.match(r'^[A-Za-z].*![A-Z]+\d+$', clean) or re.match(r'^[A-Z]+\d+$', clean):
            sheet, coord = _parse_cell_ref(expr, current_sheet)
            return _resolve_cell_ref(sheet, coord)
        # Try as number
        try:
            return float(expr) if '.' in expr else int(expr)
        except ValueError:
            return expr

    def _resolve_vlookup(lookup_expr, table_range, col_index, current_sheet):
        """Resolve VLOOKUP(lookup_val, table_range, col_index, FALSE)."""
        lookup_val = _resolve_ref_or_value(lookup_expr, current_sheet)
        if lookup_val is None:
            return None

        vl_sheet, vl_range = _parse_cell_ref(table_range, current_sheet)
        m = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', vl_range.replace('$', ''), re.IGNORECASE)
        if not m:
            return None

        start_col = _col_letter_to_num(m.group(1))
        start_row = int(m.group(2))
        end_row = int(m.group(4))

        # Search first column for lookup_val
        for r in range(start_row, end_row + 1):
            coord = f"{_num_to_col_letter(start_col)}{r}"
            val = _resolve_cell_ref(vl_sheet, coord)
            if val is not None and str(val).strip() == str(lookup_val).strip():
                # Return value from col_index column
                result_coord = f"{_num_to_col_letter(start_col + col_index - 1)}{r}"
                return _resolve_cell_ref(vl_sheet, result_coord)
        return None

    def _resolve_formula(formula, current_sheet):
        """Attempt to resolve a formula to its value."""
        if not isinstance(formula, str) or not formula.startswith('='):
            return formula

        f = formula.strip()

        # Handle =+ prefix (e.g., =+B95)
        if f.startswith('=+'):
            f = '=' + f[2:]

        # --- Simple cell reference: =A1 or ='Sheet Name'!A1 ---
        m = re.match(r"^=('?[^=+\-*/()]+?'?!)?\$?([A-Z]+)\$?(\d+)$", f, re.IGNORECASE)
        if m:
            ref = f[1:]
            sheet, coord = _parse_cell_ref(ref, current_sheet)
            return _resolve_cell_ref(sheet, coord)

        # --- =IFERROR(INDEX(...MATCH...MATCH...), "") ---
        # Most common pattern in the Report Actual sheet
        m_iferror = re.match(r'^=IFERROR\((.+),\s*""\s*\)$', f, re.IGNORECASE)
        inner = f[1:] if not m_iferror else m_iferror.group(1).strip()
        if m_iferror:
            # Resolve inner expression; if it fails, return ""
            val = _resolve_index_match(inner, current_sheet)
            return val if val is not None else ""

        # --- Direct INDEX/MATCH without IFERROR ---
        if 'INDEX(' in f.upper() and 'MATCH(' in f.upper():
            val = _resolve_index_match(f[1:] if f.startswith('=') else f, current_sheet)
            if val is not None:
                return val

        # --- =IF(X=0,"",X) or =IF(X="","",X) ---
        m = re.match(r'^=IF\(\s*(.+?)\s*=\s*(?:0|"")\s*,\s*""\s*,\s*(.+?)\s*\)$', f, re.IGNORECASE)
        if m:
            ref_str = m.group(2).strip()
            sheet, coord = _parse_cell_ref(ref_str, current_sheet)
            val = _resolve_cell_ref(sheet, coord)
            if val is None or val == 0 or val == "":
                return ""
            return val

        # --- =ROUND(X, N) ---
        m = re.match(r'^=ROUND\(\s*(.+?)\s*,\s*(\d+)\s*\)$', f, re.IGNORECASE)
        if m:
            ref_str = m.group(1).strip()
            decimals = int(m.group(2))
            sheet, coord = _parse_cell_ref(ref_str, current_sheet)
            val = _resolve_cell_ref(sheet, coord)
            if val is None:
                return None
            try:
                return round(float(val), decimals)
            except (ValueError, TypeError):
                return val

        # --- =CONCATENATE(A,B,C,...) ---
        m = re.match(r'^=CONCATENATE\((.+)\)$', f, re.IGNORECASE)
        if m:
            args_str = m.group(1)
            parts = []
            for arg in args_str.split(','):
                arg = arg.strip()
                if arg.startswith('"') and arg.endswith('"'):
                    parts.append(arg.strip('"'))
                else:
                    sheet, coord = _parse_cell_ref(arg, current_sheet)
                    val = _resolve_cell_ref(sheet, coord)
                    parts.append(str(val) if val is not None else "")
            return "".join(parts)

        # Could not resolve
        return None

    def _resolve_index_match(expr, current_sheet):
        """Resolve INDEX(range, MATCH(...), MATCH(...)) with optional /N suffix.

        Handles VLOOKUP nested inside MATCH for lookup value resolution.
        """
        # Check for trailing /100 or *100 etc.
        divisor = 1
        m_div = re.match(r'^(.+)/(\d+)\s*$', expr)
        if m_div:
            expr = m_div.group(1).strip()
            divisor = int(m_div.group(2))

        # Parse INDEX(range, MATCH(...), MATCH(...))
        # Use a more flexible parser that handles nested functions
        m = re.match(
            r"INDEX\(\s*('.+?'!\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)\s*,"
            r"\s*MATCH\((.+?),\s*('.+?'!\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)\s*,\s*0\s*\)"
            r"(?:\s*,\s*MATCH\((.+?),\s*('.+?'!\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+)\s*,\s*0\s*\))?"
            r"\s*\)",
            expr, re.IGNORECASE
        )
        if not m:
            return None

        index_range = m.group(1).strip()
        row_lookup_expr = m.group(2).strip()
        row_match_range = m.group(3).strip()
        col_lookup_expr = m.group(4)  # May be None for single-MATCH INDEX
        col_match_range = m.group(5)

        # Parse the sheet from the INDEX range
        idx_sheet, idx_range = _parse_cell_ref(index_range, current_sheet)

        # Resolve the row lookup value (may contain VLOOKUP)
        row_lookup_val = _resolve_ref_or_value(row_lookup_expr, current_sheet)
        if row_lookup_val is None:
            return None

        # Parse row MATCH range and execute
        row_match_sheet, row_match_rng = _parse_cell_ref(row_match_range, current_sheet)
        row_idx = _match_in_range(row_lookup_val, row_match_sheet, row_match_rng)
        if row_idx is None:
            return None

        if col_lookup_expr and col_match_range:
            # Two-dimensional INDEX/MATCH
            col_lookup_val = _resolve_ref_or_value(col_lookup_expr.strip(), current_sheet)
            if col_lookup_val is None:
                return None
            col_match_sheet, col_match_rng = _parse_cell_ref(col_match_range.strip(), current_sheet)
            col_idx = _match_in_range(col_lookup_val, col_match_sheet, col_match_rng)
            if col_idx is None:
                return None
        else:
            col_idx = 1  # Single column INDEX

        # Execute INDEX
        val = _get_range_value(idx_sheet, idx_range, row_idx, col_idx)
        if val is not None and divisor != 1:
            try:
                val = float(val) / divisor
            except (ValueError, TypeError):
                pass
        return val

    # Collect all formula cells that need resolution
    formula_cells = []  # [(sheet_name, coordinate, formula)]
    for sheet_name in wb_formulas.sheetnames:
        ws = wb_formulas[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    # Skip if cached value already exists
                    if (sheet_name, cell.coordinate) not in cell_values:
                        formula_cells.append((sheet_name, cell.coordinate, val))

    # Multi-pass resolution: resolve what we can, then re-try with new values
    max_passes = 3
    for pass_num in range(max_passes):
        newly_resolved = 0
        remaining = []
        for sheet_name, coord, formula in formula_cells:
            resolved = _resolve_formula(formula, sheet_name)
            if resolved is not None:
                cell_values[(sheet_name, coord)] = resolved
                newly_resolved += 1
            else:
                remaining.append((sheet_name, coord, formula))
        formula_cells = remaining
        if newly_resolved == 0:
            break  # No progress, stop

    resolved_count = sum(1 for k, v in cell_values.items() if v is not None)
    unresolved_count = len(formula_cells)
    log(f"  Formulas resolved: {resolved_count}, unresolved: {unresolved_count}")

    # Create a new clean workbook with values only
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # remove default sheet

    for sheet_name in wb_formulas.sheetnames:
        ws_src = wb_formulas[sheet_name]
        ws_dst = wb_out.create_sheet(title=sheet_name)

        for row in ws_src.iter_rows():
            for cell in row:
                dst_cell = ws_dst.cell(row=cell.row, column=cell.column)
                val = cell.value

                if isinstance(val, str) and val.startswith('='):
                    # Use resolved value from cell_values cache
                    resolved = cell_values.get((sheet_name, cell.coordinate))
                    dst_cell.value = resolved  # None if still unresolved
                else:
                    dst_cell.value = val

    # Save as temporary xlsx
    tmp_dir = file_path.parent
    resolved_path = tmp_dir / (file_path.stem + "_resolved.xlsx")
    wb_out.save(str(resolved_path))

    wb_formulas.close()
    wb_cached.close()
    wb_out.close()

    log(f"  Saved resolved file: {resolved_path.name}")
    return resolved_path


# ---------------------------------------------------------------------------
# Docling backend (primary)
# ---------------------------------------------------------------------------

def _parse_with_docling(file_path: Path, log_callback=None) -> str:
    """Parse a file to Markdown using Docling (runs locally, no API key)."""
    from docling.document_converter import DocumentConverter

    def log(msg):
        if log_callback:
            log_callback(msg)

    log(f"Parsing {file_path.name} with Docling...")

    converter = DocumentConverter()
    result = converter.convert(str(file_path))
    markdown_text = result.document.export_to_markdown()

    if not markdown_text or not markdown_text.strip():
        raise RuntimeError(f"Docling returned empty output for {file_path.name}")

    return markdown_text


# ---------------------------------------------------------------------------
# pandas/openpyxl fallback (Excel only)
# ---------------------------------------------------------------------------

def _parse_with_pandas(file_path: Path, log_callback=None) -> str:
    """Parse an Excel file to Markdown using pandas + openpyxl (local, no API)."""
    import pandas as pd

    def log(msg):
        if log_callback:
            log_callback(msg)

    log(f"Using fallback parser (pandas/openpyxl) for {file_path.name}...")

    engine = 'openpyxl'
    xls = pd.ExcelFile(file_path, engine=engine)
    md_parts = []

    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine=engine)

        if df.empty:
            continue

        md_parts.append(f"## {sheet_name}\n")

        # Try to detect a header row (first row with mostly non-null string values)
        header_row = 0
        first_row = df.iloc[0]
        non_null = first_row.dropna()
        if len(non_null) > 0 and all(isinstance(v, str) for v in non_null):
            header_row = 0
        else:
            header_row = None

        if header_row is not None:
            df.columns = [str(v) if pd.notna(v) else f"Col_{i}"
                          for i, v in enumerate(df.iloc[header_row])]
            df = df.iloc[header_row + 1:].reset_index(drop=True)

        # Convert to markdown table
        df = df.fillna("")
        df = df.astype(str)
        df = df.replace("nan", "")

        if df.empty or (df == "").all().all():
            continue

        headers = list(df.columns)
        md_parts.append("| " + " | ".join(headers) + " |")
        md_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")

        for _, row in df.iterrows():
            values = [str(v).replace("|", "\\|") for v in row]
            md_parts.append("| " + " | ".join(values) + " |")

        md_parts.append("")

    return "\n".join(md_parts)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_excel_to_markdown(file_path: Path, api_key: str = None,
                            log_callback=None) -> Path:
    """Parse an Excel file to Markdown and save alongside the original.

    For .xlsm files, formulas are resolved to values first so that
    macro-enabled workbooks display their calculated data.

    Returns the path to the generated .md file.
    """
    def log(msg):
        if log_callback:
            log_callback(msg)

    file_path = Path(file_path)
    output_path = file_path.parent / (file_path.stem + ".md")
    markdown_text = None

    # For xlsm files, resolve formulas to values first
    parse_path = file_path
    resolved_path = None
    if file_path.suffix.lower() == '.xlsm':
        try:
            resolved_path = _resolve_xlsm_formulas(file_path, log_callback)
            parse_path = resolved_path
        except Exception as e:
            log(f"Formula resolution failed: {e}. Parsing original file...")

    # Try Docling first
    try:
        markdown_text = _parse_with_docling(parse_path, log_callback)
    except Exception as e:
        log(f"Docling failed: {e}")
        log("Falling back to pandas parser...")

    # Fallback to pandas for Excel files
    if markdown_text is None and parse_path.suffix.lower() in ['.xlsx', '.xlsm']:
        markdown_text = _parse_with_pandas(parse_path, log_callback)

    # Clean up temporary resolved file
    if resolved_path and resolved_path.exists():
        try:
            resolved_path.unlink()
        except OSError:
            pass

    if markdown_text is None:
        raise RuntimeError(f"All parsers failed for {file_path.name}")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(markdown_text)
    log(f"Successfully parsed {file_path.name} -> {output_path.name}")
    return output_path


def parse_all_in_directories(directories: list[Path], api_key: str = None,
                             log_callback=None) -> list[Path]:
    """Parse all supported files in the given directories.

    Returns list of generated .md file paths.
    """
    results = []
    for input_dir in directories:
        if not input_dir.exists() or not input_dir.is_dir():
            continue
        for item in input_dir.iterdir():
            if item.is_file() and item.suffix.lower() in SUPPORTED_PARSE_EXTENSIONS:
                try:
                    md_path = parse_excel_to_markdown(item, api_key, log_callback)
                    results.append(md_path)
                except Exception:
                    continue
    return results
